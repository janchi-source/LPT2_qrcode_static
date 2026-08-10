#!/usr/bin/env python3
"""Porovná data/deti.json s Excelom. Nezávisle od generátora.

    python3 scripts/over-deti.py ~/Downloads/'skupilnky - vytlačiť (2).xlsx' \\
        --nahrada "Tobiáš Turlík=Ondrej Kullač" --nahrada "Oliver MichalÍk=Oliver Michalík"

Prepínač --nahrada je ten istý ako v generuj-deti.py: v Exceli je niekto pod
iným menom, než aké má v appke platiť. Bez neho by overovač takú výmenu hlásil
ako rozdiel.

Prečo zvlášť a nie ako súčasť generátora: generátor Excel číta, takže si vlastnú
chybu v čítaní nemá ako všimnúť. Tento skript ide na to z druhej strany —
vezme hotový výsledok a spýta sa, či v ňom je presne to, čo je v tabuľke.
Dá sa spustiť kedykoľvek, aj bez pregenerovania.

Návratový kód 0 = všetko sedí, 1 = našiel rozdiel.
"""
import json
import sys
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl

N = 10
KOREN = Path(__file__).resolve().parent.parent
chyby = []
varovania = []


def txt(x):
    return x.strip() if isinstance(x, str) else ''


def zjednodus(s):
    """Na porovnávanie mien: bez diakritiky, malé písmená, jedna medzera.

    Rozdiel „Šimon Ík" vs „Šimon ík" je preklep v Exceli, nie iné dieťa —
    nemá zmysel kvôli nemu hlásiť chybu, ale ani ho ticho prehltnúť.
    """
    s = unicodedata.normalize('NFKD', s.lower())
    return ' '.join(''.join(c for c in s if not unicodedata.combining(c)).split())


def nacitaj_excel(cesta):
    wb = openpyxl.load_workbook(cesta, data_only=True)
    podla_dvojice, harky, animatori = {}, {}, {}
    for g in range(1, N + 1):
        rows = list(wb[f'{g}.skupinka'].iter_rows(values_only=True))
        dvojica = tuple(x for x in (txt(rows[1][1]), txt(rows[1][2])) if x)
        # Bloky sa párujú podľa PRIEZVISKA prvého animátora — krstné mená sa
        # v Exceli píšu raz tak, raz onak (Ján/Janko), presné porovnanie zlyhá.
        podla_dvojice[zjednodus(dvojica[0].split()[-1])] = g
        animatori[g] = list(dvojica)
        harky[g] = [(txt(r[1]), txt(r[2])) for r in rows[2:] if txt(r[1])]

    ws = wb['skupinky pre animátorov']
    tabulka, akt = {}, {}
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        for base in (1, 6, 11):
            b = [txt(x) for x in (row[base:base + 4] + (None,) * 4)[:4]]
            if b[0] == 'Animátori:':
                g = podla_dvojice.get(zjednodus(b[1].split()[-1])) if b[1] else None
                if g is None:
                    chyby.append(f'riadok {i}, stĺpec {base + 1}: animátor „{b[1]}" '
                                 'nesedí so žiadnym hárkom po skupinkách')
                akt[base] = g
                if g:
                    tabulka.setdefault(g, [])
            elif b[0] and b[1] and akt.get(base):
                tabulka[akt[base]].append((b[0], b[1]))
    return tabulka, harky, animatori


def main():
    argv = sys.argv[1:]
    nahrady = {}
    while '--nahrada' in argv:
        i = argv.index('--nahrada')
        try:
            zle, spravne = argv[i + 1].split('=', 1)
        except (IndexError, ValueError):
            sys.exit('--nahrada čakala tvar "Meno Priezvisko=Iné Meno Priezvisko"')
        nahrady[zjednodus(zle)] = tuple(spravne.strip().split(' ', 1))
        del argv[i:i + 2]
    cesty = [a for a in argv if not a.startswith('--')]
    if not cesty:
        sys.exit('Použitie: python3 scripts/over-deti.py <cesta k xlsx> [--nahrada "A=B"]')

    tabulka, harky, animatori = nacitaj_excel(cesty[0])
    if nahrady:
        for g in tabulka:
            tabulka[g] = [nahrady.get(zjednodus(' '.join(x)), x) for x in tabulka[g]]
        for g in harky:
            harky[g] = [nahrady.get(zjednodus(' '.join(x)), x) for x in harky[g]]
        for g in animatori:
            animatori[g] = [' '.join(nahrady[zjednodus(x)]) if zjednodus(x) in nahrady else x
                            for x in animatori[g]]
    data = json.loads((KOREN / 'data' / 'deti.json').read_text(encoding='utf-8'))
    deti = data['deti']

    print(f'Excel (tabuľka): {sum(len(v) for v in tabulka.values())} detí')
    print(f'data/deti.json : {len(deti)} detí\n')

    # --- 1. dieťa po dieťati, skupinka po skupinke --------------------------
    for g in range(1, N + 1):
        v_tabulke = sorted(tabulka.get(g, []), key=lambda x: (zjednodus(x[1]), zjednodus(x[0])))
        v_datach = sorted(((d['meno'], d['priezvisko']) for d in deti if d['skupina'] == g),
                          key=lambda x: (zjednodus(x[1]), zjednodus(x[0])))
        if len(v_tabulke) != len(v_datach):
            chyby.append(f'sk.{g}: tabuľka má {len(v_tabulke)} detí, dáta {len(v_datach)}')
        t = [zjednodus(' '.join(x)) for x in v_tabulke]
        d = [zjednodus(' '.join(x)) for x in v_datach]
        for x in set(t) - set(d):
            chyby.append(f'sk.{g}: „{x}" je v Exceli, ale nie v dátach')
        for x in set(d) - set(t):
            chyby.append(f'sk.{g}: „{x}" je v dátach, ale nie v Exceli')
        # presné znenie mena (diakritika, veľké písmená)
        for a, b in zip(v_tabulke, v_datach):
            if a != b and zjednodus(' '.join(a)) == zjednodus(' '.join(b)):
                varovania.append(f'sk.{g}: „{" ".join(b)}" v dátach vs „{" ".join(a)}" v Exceli')

    # --- 2. animátori --------------------------------------------------------
    for g in range(1, N + 1):
        v_datach = data['animatori'].get(str(g)) or []
        if [zjednodus(x) for x in v_datach] != [zjednodus(x) for x in animatori[g]]:
            chyby.append(f'sk.{g}: animátori {v_datach} != {animatori[g]}')
        elif v_datach != animatori[g]:
            varovania.append(f'sk.{g}: animátori v dátach {v_datach} vs v Exceli {animatori[g]}')

    # --- 3. duplicity a úplnosť ---------------------------------------------
    mena = Counter(zjednodus(f'{d["meno"]} {d["priezvisko"]}') for d in deti)
    for m, n in mena.items():
        if n > 1:
            chyby.append(f'dieťa „{m}" je v dátach {n}×')

    naramky = [d['naramok'] for d in deti]
    if len(set(naramky)) != len(naramky):
        chyby.append('dve deti majú to isté číslo náramku')
    # Súvislý rad sa nevyžaduje — deti pridané po vytlačení náramkov dostávajú
    # ďalšie voľné čísla. Diery sú v poriadku, duplicity nie.
    diery = sorted(set(range(1, max(naramky) + 1)) - set(naramky))
    if diery:
        varovania.append(f'nevyužité čísla náramkov: {diery}')
    if len({d['id'] for d in deti}) != len(deti):
        chyby.append('duplicitné ID (QR kódy)')

    # --- 4. rozdiel medzi tabuľkou a hárkami (len na vedomie) ---------------
    for g in range(1, N + 1):
        t = {zjednodus(' '.join(x)) for x in tabulka.get(g, [])}
        h = {zjednodus(' '.join(x)) for x in harky[g]}
        for x in t - h:
            varovania.append(f'sk.{g}: „{x}" je v tabuľke, ale nie v hárku {g}.skupinka')
        for x in h - t:
            varovania.append(f'sk.{g}: „{x}" je v hárku {g}.skupinka, ale nie v tabuľke')

    # --- výsledok ------------------------------------------------------------
    if varovania:
        print('⚠️  Na vedomie (nie je to chyba dát appky):')
        for v in varovania:
            print('   ', v)
        print()
    if chyby:
        print('❌ NESEDÍ:')
        for c in chyby:
            print('   ', c)
        sys.exit(1)

    print('✅ Všetkých %d detí sedí s Excelom — mená, skupinky aj animátori.' % len(deti))
    print('   Veľkosti skupiniek:', [len(tabulka[g]) for g in range(1, N + 1)])


if __name__ == '__main__':
    main()
