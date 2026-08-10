#!/usr/bin/env python3
"""Vygeneruje data/deti.json z Excelu — PRÍRASTKOVO.

    pip3 install openpyxl
    python3 scripts/generuj-deti.py ~/Downloads/'skupilnky - vytlačiť (2).xlsx'
    python3 scripts/over-deti.py   ~/Downloads/'skupilnky - vytlačiť (2).xlsx'
    npm test

DÔLEŽITÉ: náramky sú vytlačené a nalepené, takže sa NESMÚ prečíslovať. Skript
preto berie existujúci data/deti.json ako základ a deťom, ktoré v ňom už sú,
NEMENÍ nič — ani číslo náramku, ani QR kód, ani skupinku, ani trasu. Nové deti
len pridá s ďalšími voľnými číslami na konci.

Keď Excel presunie už existujúce dieťa do inej skupinky, skript SKONČÍ CHYBOU —
taký presun by mu zmenil trasu a vytlačený náramok by prestal sedieť. Ak sa to
naozaj má stať, treba to povoliť prepínačom --povol-presun (a vytlačiť nanovo).

Prepínače:
  --nahrada "Meno Priezvisko=Iné Meno"   v Exceli je niekto pod iným menom;
                                         použije sa meno vpravo (dá sa opakovať)
  --vymen "Kto odchádza=Kto prichádza"   na náramku pôvodného dieťaťa bude
                                         odteraz niekto iný. Číslo, QR kód,
                                         skupinka aj trasa ostávajú — mení sa
                                         len meno, takže stačí vytlačiť ten
                                         jeden štítok nanovo.
  --povol-presun                         dovolí presun dieťaťa do inej skupinky
  --od-nuly                              zahodí doterajšie priradenie a rozdá
                                         čísla nanovo (LEN pred prvou tlačou!)

Čo súbor obsahuje a prečo:

  skupina  — číslo skupinky z Excelu (1–10). Určuje, kde dieťa skončí.
  trieda   — ktorá z piatich možných trás (0–4, podľa počtu prevedení
             0/2/4/6/8). Šablóny sú v lib/hra.js; samotná trasa sa
             z (skupina, trieda) dopočíta tam, tu sa neukladá.
  naramok  — číslo náramku. Pri prvom rozdaní sa čísla prideľujú v BLOKOCH
             podľa štartového stanovišťa, aby sa deti dali na začiatku rozdeliť
             podľa čísel (manuál, s. 2). Deti pridané neskôr dostanú ďalšie
             voľné čísla — tie do blokov nezapadnú a v rozpise sú vypísané
             zvlášť.

Každá skupinka rozdelí deti medzi trasy rovnako (POCTY nižšie). Vďaka tomu je
na každom stanovišti v každom kole presne 10 alebo 11 detí — overuje
test/trasy.test.js.
"""
import json
import random
import sys
import unicodedata
from pathlib import Path

import openpyxl

N = 10
TRIEDA_NAVYSE = 10        # šablóna pre 11. dieťa v skupinke
SEED_TRIED = 1000         # poradie tried vnútri skupinky
SEED_NARAMKOV = 20260810  # zamiešanie čísel vnútri štartového bloku

KOREN = Path(__file__).resolve().parent.parent

# Musí sedieť so SABLONY v lib/hra.js — test/trasy.test.js to porovná.
# Päť trás, ktoré pravidlá manuálu dovoľujú (0, 2, 4, 6 a 8 prevedení).
SABLONY = [
    [1, 2, 3, 4, 5, 6, 7, 8, 9, 0],
    [9, 1, 2, 3, 4, 5, 6, 7, 8, 0],
    [7, 9, 1, 2, 3, 4, 5, 6, 8, 0],
    [5, 7, 9, 1, 2, 3, 4, 6, 8, 0],
    [3, 5, 7, 9, 1, 2, 4, 6, 8, 0],
]

# Koľko detí v skupinke ide ktorou trasou. Súčet = 10; skupinka s 11 deťmi
# pridá jedno navyše na trasu TRIEDA_NAVYSE.
#
# Toto je jediné, čo sa pri trasách dá voliť — a určuje, koľko detí sa
# v ktorom kole na stanovišti obmení:
#
#   kolo      1  2  3  4  5  6  7  8  9
#   obmení sa 9  7  5  3  0  3  5  7  9   detí z 10
#
# Nula v 5. kole je vlastnosť pravidiel, nie tohto nastavenia: piaty krok je
# +1 vo všetkých piatich trasách (viď lib/hra.js).
POCTY = [1, 2, 2, 2, 3]
TRIEDA_NAVYSE = 2


def domovske(skupina):
    """Štvrť skupinky — tam, kde ju manuál (s. 4) necháva v poslednom kole."""
    return (N - skupina) % N


def trasa(skupina, trieda):
    H = domovske(skupina)
    return [(H + p) % N for p in SABLONY[trieda]]


def prezdravie(x):
    return x.strip() if isinstance(x, str) else ''


def kluc(meno, priezvisko):
    """Kľúč na párovanie detí medzi Excelom a doterajšími dátami.

    Bez diakritiky a bez veľkých písmen: v Exceli sa to isté dieťa píše raz
    „Šimon Ík", raz „Šimon ík", raz „FILIPOVA". Keby sa párovalo presne,
    appka by ho považovala za nové a pridelila mu druhý náramok.
    """
    s = unicodedata.normalize('NFKD', f'{meno} {priezvisko}'.lower())
    return ' '.join(''.join(c for c in s if not unicodedata.combining(c)).split())


def nacitaj(cesta):
    """Prečíta deti z hárku „skupinky pre animátorov".

    Zdrojom pravdy je ZÁMERNE prehľadová tabuľka, nie hárky 1.skupinka…
    10.skupinka. Tie sa v praxi opravujú menej dôsledne — v auguste 2026 v nich
    chýbalo jedno dieťa a jedno bolo pod cudzím menom, kým tabuľka bola
    v poriadku. Hárky sa preto načítajú tiež, ale len na porovnanie: rozdiely
    sa vypíšu ako upozornenie, nech sa o nich vie.

    Tabuľka má tri stĺpcové bloky (1, 6, 11) a niekoľko pásiem riadkov.
    Pásmo sa začína riadkom „Animátori:", podľa ktorého sa pozná, o ktorú
    skupinku ide — dvojica animátorov je kľúč do hárkov po skupinkách.
    """
    wb = openpyxl.load_workbook(cesta, data_only=True)

    # 1) hárky po skupinkách: dvojica animátorov → číslo skupinky + zoznam detí
    animatori, harky = {}, {}
    podla_dvojice = {}
    for g in range(1, N + 1):
        nazov = f'{g}.skupinka'
        if nazov not in wb.sheetnames:
            sys.exit(f'V súbore chýba hárok „{nazov}".')
        rows = list(wb[nazov].iter_rows(values_only=True))
        dvojica = tuple(x for x in (prezdravie(rows[1][1]), prezdravie(rows[1][2])) if x)
        if len(dvojica) < 2:
            sys.exit(f'Hárok „{nazov}" nemá v druhom riadku dvojicu animátorov.')
        # Bloky sa párujú podľa PRIEZVISKA prvého animátora — krstné mená sa
        # v súbore píšu raz tak, raz onak (Ján/Janko, Lívia/Línia) a presné
        # porovnanie by zlyhalo.
        priezvisko = kluc('', dvojica[0].split()[-1])
        if priezvisko in podla_dvojice:
            sys.exit(f'Priezvisko animátora „{dvojica[0]}" je v súbore dvakrát — '
                     'podľa neho sa priraďujú bloky v prehľadovej tabuľke.')
        animatori[g] = list(dvojica)
        podla_dvojice[priezvisko] = g
        harky[g] = [(prezdravie(r[1]), prezdravie(r[2])) for r in rows[2:] if prezdravie(r[1])]

    # 2) prehľadová tabuľka
    if 'skupinky pre animátorov' not in wb.sheetnames:
        sys.exit('V súbore chýba hárok „skupinky pre animátorov".')
    ws = wb['skupinky pre animátorov']
    tabulka = {}
    aktualna = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        for base in (1, 6, 11):
            b = [prezdravie(x) for x in (row[base:base + 4] + (None,) * 4)[:4]]
            if b[0] == 'Animátori:':
                g = podla_dvojice.get(kluc('', b[1].split()[-1])) if b[1] else None
                if g is None:
                    sys.exit(f'Riadok {i}, stĺpec {base + 1}: animátor „{b[1]}" '
                             'nesedí so žiadnym hárkom po skupinkách. Skontroluj mená.')
                aktualna[base] = g
                tabulka.setdefault(g, [])
            elif b[0] and b[1] and aktualna.get(base):
                tabulka[aktualna[base]].append({'meno': b[0], 'priezvisko': b[1],
                                                'rocnik': b[2] or None})

    chyba = [g for g in range(1, N + 1) if g not in tabulka]
    if chyba:
        sys.exit(f'V prehľadovej tabuľke sa nenašli skupinky: {chyba}')

    # 3) kontroly, ktoré musia prejsť
    videne = {}
    for g in tabulka:
        for d in tabulka[g]:
            k = kluc(d['meno'], d['priezvisko'])
            if k in videne:
                sys.exit(f'Dieťa „{d["meno"]} {d["priezvisko"]}" je v tabuľke dvakrát '
                         f'(skupinka {videne[k]} aj {g}). Oprav to v Exceli.')
            videne[k] = g

    # 4) porovnanie s hárkami — len upozornenie
    rozdiely = []
    for g in range(1, N + 1):
        t = [(d['meno'], d['priezvisko']) for d in tabulka[g]]
        h = harky[g]
        hk = {kluc(*x) for x in h}
        tk = {kluc(*x) for x in t}
        chyba_v_harku = [x for x in t if kluc(*x) not in hk]
        navyse_v_harku = [x for x in h if kluc(*x) not in tk]
        if chyba_v_harku or navyse_v_harku:
            rozdiely.append((g, chyba_v_harku, navyse_v_harku))
    if rozdiely:
        print('⚠️  Prehľadová tabuľka a hárky po skupinkách sa nezhodujú.')
        print('    Použila sa TABUĽKA. Rozdiely (nech sa o nich vie):')
        for g, chyba_v, navyse in rozdiely:
            for x in chyba_v:
                print(f'      sk.{g}: „{" ".join(x)}" je v tabuľke, ale nie v hárku {g}.skupinka')
            for x in navyse:
                print(f'      sk.{g}: „{" ".join(x)}" je v hárku {g}.skupinka, ale nie v tabuľke')
        print()

    return [{'cislo': g, 'animatori': animatori[g], 'deti': tabulka[g]} for g in range(1, N + 1)]


def nacitaj_doterajsie():
    """Doterajšie priradenie náramkov. Prázdne, keď sa generuje prvýkrát."""
    subor = KOREN / 'data' / 'deti.json'
    if not subor.exists():
        return {}
    data = json.loads(subor.read_text(encoding='utf-8'))
    return {kluc(d['meno'], d['priezvisko']): d for d in data.get('deti', [])}


def volna_trieda(profil, velkost):
    """Ktorou trasou má ísť ďalšie dieťa v skupinke.

    Profil skupinky musí ostať [1,2,2,2,3] (10 detí) alebo [1,2,3,2,3] (11).
    Vďaka tomu je na každom stanovišti v každom kole 10 alebo 11 detí — všetky
    skupinky delia deti medzi trasy rovnako, len tá jedenásta ide navyše po
    trase TRIEDA_NAVYSE. Keď by skupinka mala 12, invariant sa poruší a treba
    to vedieť.
    """
    if velkost == sum(POCTY):          # z 10 na 11 → jedenásta trasa
        return TRIEDA_NAVYSE, None
    # 12 a viac: pridá sa na najmenej obsadenú trasu, ale je to na vedomie
    najmenej = min(range(len(POCTY)), key=lambda t: profil[t])
    return najmenej, (f'skupinka bude mať {velkost + 1} detí — na jednom '
                      f'stanovišti tak môže byť 12 detí namiesto 10–11')


def main():
    argv = sys.argv[1:]
    nahrady = {}
    vymeny = {}
    povol_presun = '--povol-presun' in argv
    od_nuly = '--od-nuly' in argv
    while '--nahrada' in argv:
        i = argv.index('--nahrada')
        try:
            zle, spravne = argv[i + 1].split('=', 1)
        except (IndexError, ValueError):
            sys.exit('--nahrada čakala tvar "Meno Priezvisko=Iné Meno Priezvisko"')
        nahrady[kluc(*zle.strip().split(' ', 1))] = spravne.strip().split(' ', 1)
        del argv[i:i + 2]
    while '--vymen' in argv:
        i = argv.index('--vymen')
        try:
            odchadza, prichadza = argv[i + 1].split('=', 1)
        except (IndexError, ValueError):
            sys.exit('--vymen čakala tvar "Kto odchádza=Kto prichádza"')
        vymeny[kluc(*odchadza.strip().split(' ', 1))] = prichadza.strip().split(' ', 1)
        del argv[i:i + 2]
    cesty = [a for a in argv if not a.startswith('--')]
    if not cesty:
        sys.exit('Použitie: python3 scripts/generuj-deti.py <cesta k xlsx> [prepínače]')

    skupiny = nacitaj(cesty[0])

    # --- náhrady mien --------------------------------------------------------
    # Platia aj na animátorov: v Exceli sa raz píše „Lívia", raz „Línia",
    # a v rozpise má byť to správne meno.
    if nahrady:
        pouzite = set()
        for s in skupiny:
            for i, meno_anim in enumerate(s['animatori']):
                k = kluc(*(meno_anim.split(' ', 1) + [''])[:2])
                if k in nahrady:
                    s['animatori'][i] = ' '.join(nahrady[k]).strip()
                    print(f"  náhrada: sk.{s['cislo']} animátor \u201e{meno_anim}\u201c"
                          f" \u2192 \u201e{s['animatori'][i]}\u201c")
                    pouzite.add(k)
            for d in s['deti']:
                k = kluc(d['meno'], d['priezvisko'])
                if k in nahrady:
                    meno, priezvisko = (nahrady[k] + [''])[:2]
                    print(f"  náhrada: sk.{s['cislo']} \u201e{d['meno']} {d['priezvisko']}\u201c"
                          f" \u2192 \u201e{meno} {priezvisko}\u201c")
                    d['meno'], d['priezvisko'] = meno, priezvisko
                    pouzite.add(k)
        for k in set(nahrady) - pouzite:
            sys.exit(f'--nahrada: „{k}" sa v Exceli nenašlo, takže sa nič nenahradilo.')
        print()

    doterajsie = {} if od_nuly else nacitaj_doterajsie()

    # --- výmena na náramku ---------------------------------------------------
    # Dieťa nepríde a jeho miesto zaberie iné. Náramok, QR kód, skupinka aj
    # trasa ostávajú — nové dieťa ich zdedí. Bez tohto by sa ten, čo odišiel,
    # zmazal a nový by dostal ďalšie voľné číslo, čo je zbytočná zmena.
    for stary_kluc, novy in vymeny.items():
        zaznam = doterajsie.pop(stary_kluc, None)
        if zaznam is None:
            sys.exit(f'--vymen: „{stary_kluc}" sa v doterajších dátach nenašlo.')
        meno, priezvisko = (novy + [''])[:2]
        print(f"  výmena: náramok {zaznam['naramok']} (sk.{zaznam['skupina']}) "
              f"\u201e{zaznam['meno']} {zaznam['priezvisko']}\u201c "
              f"\u2192 \u201e{meno} {priezvisko}\u201c "
              f"\u2014 vytlač preň nový štítok")
        zaznam = {**zaznam, 'meno': meno, 'priezvisko': priezvisko}
        doterajsie[kluc(meno, priezvisko)] = zaznam
    if vymeny:
        print()
    if od_nuly and (KOREN / 'data' / 'deti.json').exists():
        print('⚠️  --od-nuly: doterajšie čísla náramkov sa zahadzujú a rozdávajú nanovo.')
        print('    Ak sú náramky už vytlačené, TOTO ICH ZNEPLATNÍ.\n')

    # --- roztriedenie: čo ostáva, čo pribúda, čo odchádza --------------------
    z_excelu = {}
    for s in skupiny:
        for d in s['deti']:
            z_excelu[kluc(d['meno'], d['priezvisko'])] = (s['cislo'], d)

    ostavaju, pribudli = [], []
    presuny = []
    for k, (g, d) in z_excelu.items():
        stare = doterajsie.get(k)
        if stare is None:
            pribudli.append((g, d))
        else:
            if stare['skupina'] != g:
                presuny.append((f"{d['meno']} {d['priezvisko']}", stare['skupina'], g))
            ostavaju.append((g, d, stare))
    odisli = [v for k, v in doterajsie.items() if k not in z_excelu]

    if presuny and not povol_presun:
        print('❌ Excel presúva už existujúce deti do inej skupinky. Tým by sa im')
        print('   zmenila trasa a vytlačený náramok by prestal sedieť:')
        for meno, zo, do in presuny:
            print(f'      {meno}: sk.{zo} → sk.{do}')
        sys.exit('   Ak to naozaj tak má byť, spusti znova s --povol-presun.')

    # --- zostavenie výsledku -------------------------------------------------
    vysledok = []
    for g, d, stare in ostavaju:
        vysledok.append({**stare, 'meno': d['meno'], 'priezvisko': d['priezvisko'],
                         'rocnik': d.get('rocnik'), 'skupina': g,
                         'trieda': stare['trieda'] if not povol_presun or stare['skupina'] == g
                         else stare['trieda']})

    if pribudli:
        profily = {}
        for z in vysledok:
            profily.setdefault(z['skupina'], [0] * len(POCTY))[z['trieda']] += 1
        dalsi = max([z['naramok'] for z in vysledok], default=0) + 1
        for g, d in sorted(pribudli, key=lambda x: (x[0], x[1]['priezvisko'])):
            profil = profily.setdefault(g, [0] * len(POCTY))
            trieda, varovanie = volna_trieda(profil, sum(profil))
            if varovanie:
                print(f'⚠️  sk.{g}: {varovanie}')
            profil[trieda] += 1
            vysledok.append({'id': 'D%03d' % dalsi, 'naramok': dalsi,
                             'meno': d['meno'], 'priezvisko': d['priezvisko'],
                             'rocnik': d.get('rocnik'), 'skupina': g, 'trieda': trieda})
            print(f"➕ pridané: {d['meno']} {d['priezvisko']} → sk.{g}, "
                  f'náramok {dalsi}, štart {miesta(g, trieda)}')
            dalsi += 1
        print()

    if odisli:
        print('➖ v Exceli už nie sú (vypadli zo zoznamu):')
        for z in odisli:
            print(f"      {z['meno']} {z['priezvisko']} (náramok {z['naramok']}, sk.{z['skupina']})")
        print('   Ich čísla ostávajú voľné, ostatným sa nič nemení.\n')

    # --- prvé rozdanie: čísla v blokoch podľa štartového stanovišťa ----------
    if not doterajsie:
        rng = random.Random(SEED_NARAMKOV)
        podla_startu = {}
        for z in vysledok:
            podla_startu.setdefault(trasa(z['skupina'], z['trieda'])[0], []).append(z)
        cislo = 1
        for st in range(N):
            v_bloku = podla_startu.get(st, [])
            rng.shuffle(v_bloku)
            for z in v_bloku:
                z['naramok'] = cislo
                z['id'] = 'D%03d' % cislo
                cislo += 1

    vysledok.sort(key=lambda z: z['naramok'])

    # --- štartové bloky sa počítajú zo skutočnosti --------------------------
    bloky = []
    for st in range(N):
        cisla = sorted(z['naramok'] for z in vysledok if trasa(z['skupina'], z['trieda'])[0] == st)
        zac = None
        for i, c in enumerate(cisla):
            if zac is None:
                zac = c
            if i + 1 == len(cisla) or cisla[i + 1] != c + 1:
                bloky.append({'stanoviste': st, 'od': zac, 'do': c})
                zac = None

    vystup = {
        'deti': vysledok,
        'animatori': {str(s['cislo']): s['animatori'] for s in skupiny},
        'startove_bloky': bloky,
    }
    cielovy = KOREN / 'data' / 'deti.json'
    cielovy.write_text(json.dumps(vystup, ensure_ascii=False, indent=1), encoding='utf-8')
    print(f'{len(vysledok)} detí → {cielovy}')
    for b in bloky:
        koniec = f"–{b['do']}" if b['do'] != b['od'] else '   '
        print(f"  náramky {b['od']:>3}{koniec:<4} → stanovište {b['stanoviste']} "
              f"({b['do'] - b['od'] + 1} detí)")
    print('\nTeraz spusti `npm test` — overí, že hra s novými dátami vyjde.')


def miesta(skupina, trieda):
    return f'stanovište {trasa(skupina, trieda)[0]}'


if __name__ == '__main__':
    main()
